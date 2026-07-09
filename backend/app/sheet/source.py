"""Locate and load the Master Cultivation Reference workbook.

The sheet is the single source of truth. It lives in the operator's synced
"Mushrooms" Google Drive folder as a real .xlsx. This module returns an openpyxl
workbook from whichever source is configured, in priority order:

  1. ``--path`` / ``MASTER_SHEET_PATH``  — a local .xlsx (synced Drive folder,
     a checked-out fixture, or a manual download). Best for offline / CI.
  2. Google Drive media download of ``MASTER_SHEET_FILE_ID`` using either a
     service-account key (``GOOGLE_SERVICE_ACCOUNT_JSON``, recommended — never
     expires, set once) or a short-lived OAuth access token
     (``GOOGLE_OAUTH_TOKEN``). Best for an unattended / button-triggered sync.

Keeping the fetch behind one function means the parser and sinks never care
where the bytes came from.
"""
from __future__ import annotations

import io
import json
import os

import httpx
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

_DRIVE_SCOPE = "https://www.googleapis.com/auth/drive.readonly"
# Read+write scopes for the reverse (App -> Sheet) direction. Full ``drive``
# (not ``drive.file``, which only covers files the app itself created) so the
# documented flow — share a *pre-existing* workbook with the service account —
# actually works for the Drive .xlsx upload; ``spreadsheets`` powers the native
# Sheets API.
_WRITE_SCOPES = (
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
)

# Native Google-Sheets MIME type (vs. a real .xlsx binary stored on Drive).
GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# The canonical file in Isaac's Drive ("Master Cultivation Reference.xlsx").
DEFAULT_FILE_ID = "1KJSAauzZ-CBpA1f4hDISsLzAiFnoh4jC"

_DRIVE_MEDIA_URL = "https://www.googleapis.com/drive/v3/files/{id}"
_DRIVE_EXPORT_URL = "https://www.googleapis.com/drive/v3/files/{id}/export"
_DRIVE_UPLOAD_URL = "https://www.googleapis.com/upload/drive/v3/files/{id}"


def load_bytes_from_drive(file_id: str, token: str, *, timeout: float = 30.0) -> bytes:
    """Download an .xlsx from Google Drive via the v3 media endpoint."""
    resp = httpx.get(
        _DRIVE_MEDIA_URL.format(id=file_id),
        params={"alt": "media", "supportsAllDrives": "true"},
        headers={"Authorization": f"Bearer {token}"},
        timeout=timeout,
        follow_redirects=True,
    )
    resp.raise_for_status()
    return resp.content


def export_google_sheet(file_id: str, token: str, *, timeout: float = 30.0) -> bytes:
    """Export a *native* Google Sheet as .xlsx bytes.

    A Google Sheet has no binary to media-download; Drive exports it on the fly.
    This is what lets a Google Sheet — not just an uploaded .xlsx — be the
    source of truth.
    """
    resp = httpx.get(
        _DRIVE_EXPORT_URL.format(id=file_id),
        params={"mimeType": XLSX_MIME},
        headers={"Authorization": f"Bearer {token}"},
        timeout=timeout,
        follow_redirects=True,
    )
    resp.raise_for_status()
    return resp.content


def drive_mime_type(file_id: str, token: str, *, timeout: float = 30.0) -> str:
    """Return a Drive file's MIME type so callers can pick media vs. export."""
    resp = httpx.get(
        _DRIVE_MEDIA_URL.format(id=file_id),
        params={"fields": "mimeType", "supportsAllDrives": "true"},
        headers={"Authorization": f"Bearer {token}"},
        timeout=timeout,
        follow_redirects=True,
    )
    resp.raise_for_status()
    return resp.json().get("mimeType", "")


def load_workbook_bytes(file_id: str, token: str, *, timeout: float = 30.0) -> bytes:
    """.xlsx bytes for any Drive file — exports native Sheets, media-downloads
    a real .xlsx — so the reader doesn't care which kind it is."""
    if drive_mime_type(file_id, token, timeout=timeout) == GOOGLE_SHEET_MIME:
        return export_google_sheet(file_id, token, timeout=timeout)
    return load_bytes_from_drive(file_id, token, timeout=timeout)


def upload_xlsx_to_drive(file_id: str, data: bytes, token: str, *,
                         timeout: float = 60.0) -> None:
    """Overwrite an existing Drive .xlsx file's contents (media update).

    Used by the write-back path when the source of truth is a binary .xlsx on
    Drive rather than a native Google Sheet.
    """
    resp = httpx.patch(
        _DRIVE_UPLOAD_URL.format(id=file_id),
        params={"uploadType": "media", "supportsAllDrives": "true"},
        headers={"Authorization": f"Bearer {token}", "Content-Type": XLSX_MIME},
        content=data,
        timeout=timeout,
    )
    resp.raise_for_status()


def _service_account_token(raw_json: str, scopes=(_DRIVE_SCOPE,)) -> str:
    """Mint a Google access token from a service-account key JSON."""
    # Imported lazily so the local/path workflow needs no Google libraries.
    from google.oauth2 import service_account
    from google.auth.transport.requests import Request

    creds = service_account.Credentials.from_service_account_info(
        json.loads(raw_json), scopes=list(scopes)
    )
    creds.refresh(Request())
    return creds.token


def resolve_token(scopes=(_DRIVE_SCOPE,), *, token: str | None = None) -> str | None:
    """Best available Google access token for ``scopes`` from the environment.

    Prefers a service-account key (``GOOGLE_SERVICE_ACCOUNT_JSON``, never
    expires); falls back to a short-lived OAuth token. An explicit ``token``
    wins outright. Returns None when nothing is configured.
    """
    if token:
        return token
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if sa_json:
        return _service_account_token(sa_json, scopes)
    return os.environ.get("GOOGLE_OAUTH_TOKEN") or os.environ.get("GOOGLE_ACCESS_TOKEN")


def resolve_write_token(*, token: str | None = None) -> str | None:
    """A token with read+write scope for the App -> Sheet direction."""
    return resolve_token(_WRITE_SCOPES, token=token)


def credentials_available() -> bool:
    """Whether *some* Google credential is configured — without minting a
    token. Callers reporting status must use this rather than resolve_token,
    which does a network refresh for a service-account key."""
    return bool(
        os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
        or os.environ.get("GOOGLE_OAUTH_TOKEN")
        or os.environ.get("GOOGLE_ACCESS_TOKEN")
    )


def resolve_read_file_id(file_id: str | None = None) -> str:
    """The Drive file id the read path pulls from, in precedence order. A
    native Google Sheet (``MASTER_SHEET_GOOGLE_ID``) counts as a source —
    ``load_workbook_bytes`` exports it — so pull and status agree on the file."""
    return (
        file_id
        or os.environ.get("MASTER_SHEET_FILE_ID")
        or os.environ.get("MASTER_SHEET_GOOGLE_ID")
        or DEFAULT_FILE_ID
    )


def describe_source() -> dict:
    """Describe the configured read source without doing any I/O. Single source
    of truth for both ``resolve_workbook`` and the /api/sync/status endpoint."""
    path = os.environ.get("MASTER_SHEET_PATH")
    if path:
        return {"configured": True, "kind": "local_xlsx", "ref": path}
    for kind, var in (("google_sheet", "MASTER_SHEET_GOOGLE_ID"),
                      ("drive_xlsx", "MASTER_SHEET_FILE_ID")):
        ref = os.environ.get(var)
        if ref:
            return {"configured": True, "kind": kind, "ref": ref}
    return {"configured": False, "kind": None, "ref": None}


def resolve_workbook(path: str | None = None, file_id: str | None = None,
                     token: str | None = None) -> Workbook:
    """Return a read-only openpyxl workbook from the configured source.

    Explicit arguments win; otherwise the environment is consulted. Raises
    RuntimeError with an actionable message when nothing is configured.
    """
    path = path or os.environ.get("MASTER_SHEET_PATH")
    if path:
        return load_workbook(path, read_only=True, data_only=True)

    file_id = resolve_read_file_id(file_id)
    token = resolve_token(token=token)
    if file_id and token:
        # load_workbook_bytes handles both a native Google Sheet (exported) and
        # a real .xlsx (media-downloaded).
        data = load_workbook_bytes(file_id, token)
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    raise RuntimeError(
        "No workbook source configured. Set MASTER_SHEET_PATH to a local .xlsx, "
        "or GOOGLE_SERVICE_ACCOUNT_JSON / GOOGLE_OAUTH_TOKEN (+ optional "
        "MASTER_SHEET_FILE_ID / MASTER_SHEET_GOOGLE_ID) to pull it from Google Drive."
    )
