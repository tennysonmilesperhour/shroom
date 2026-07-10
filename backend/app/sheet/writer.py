"""Sinks for the reverse direction: write app data *into* a workbook/sheet.

Three interchangeable backends implement the same tiny interface so the
exporter (``export.push``) and the auto-push (``autosync``) don't care where
the bytes end up:

* :class:`XlsxWriter`        — a local .xlsx (openpyxl). Offline, fully tested,
  and the engine the download-workbook endpoint uses.
* :class:`DriveXlsxWriter`   — an .xlsx stored on Google Drive: download the
  current file, mutate it, re-upload. Two-way sync for a binary workbook.
* :class:`GoogleSheetsWriter`— a *native* Google Sheet via the Sheets API v4.
  Cell-level live updates; the recommended source of truth.

Interface (see :class:`SheetWriter`):
    replace_tab(tab, header, rows)  — overwrite a whole worksheet
    append_row(tab, header, row)    — add one row (creating the tab+header if
                                      it doesn't exist yet)
    commit()                        — flush pending changes to the destination

``resolve_writer`` picks a backend from the environment, mirroring how
``source.resolve_workbook`` picks a read source.
"""
from __future__ import annotations

import io
import os
import urllib.parse
from typing import Protocol, runtime_checkable

from openpyxl import Workbook, load_workbook

from . import layout, source


@runtime_checkable
class SheetWriter(Protocol):
    def replace_tab(self, tab: str, header: list[str], rows: list[list]) -> None: ...
    def upsert_tab(self, tab: str, header: list[str], rows: list[list],
                   key_cols: tuple[str, ...]) -> dict: ...
    def append_row(self, tab: str, header: list[str], row: list) -> None: ...
    def commit(self) -> None: ...
    def close(self) -> None: ...


# --------------------------------------------------------------------------- #
# Header-matching helpers (shared by every backend's upsert)
# --------------------------------------------------------------------------- #
def _norm(value: object) -> str:
    return str(value).strip().lower() if value not in (None, "") else ""


def _match_col(labels_norm: list[str], target: str) -> int:
    """Index of ``target`` in a header row: exact (normalized) match first, then
    a substring match so "Flush" finds "Flush #" and "Tub" finds "Tub ID".
    Returns -1 when absent (the caller then allocates a new column)."""
    t = _norm(target)
    if not t:
        return -1
    for i, label in enumerate(labels_norm):
        if label == t:
            return i
    for i, label in enumerate(labels_norm):
        if label and (t in label or label in t):
            return i
    return -1


def _col_letter(idx0: int) -> str:
    """0-based column index -> A1 column letters (0->A, 26->AA)."""
    idx0 += 1
    out = ""
    while idx0:
        idx0, rem = divmod(idx0 - 1, 26)
        out = chr(65 + rem) + out
    return out


# --------------------------------------------------------------------------- #
# openpyxl-backed writers (local .xlsx and Drive .xlsx)
# --------------------------------------------------------------------------- #
class _WorkbookWriter:
    """Shared openpyxl mechanics. Subclasses supply load/save of the bytes."""

    def __init__(self, wb: Workbook):
        self.wb = wb

    @staticmethod
    def _new_or_load(data: bytes | None) -> Workbook:
        if data:
            return load_workbook(io.BytesIO(data))
        wb = Workbook()
        # Drop the default empty "Sheet" so only tabs we write remain.
        wb.remove(wb.active)
        return wb

    def _sheet(self, tab: str, header: list[str], *, create: bool):
        if tab in self.wb.sheetnames:
            return self.wb[tab]
        if not create:
            return None
        ws = self.wb.create_sheet(title=tab)
        ws.append(header)
        return ws

    def replace_tab(self, tab: str, header: list[str], rows: list[list]) -> None:
        if tab in self.wb.sheetnames:
            self.wb.remove(self.wb[tab])
        ws = self.wb.create_sheet(title=tab)
        ws.append(header)
        for row in rows:
            ws.append(list(row))

    def append_row(self, tab: str, header: list[str], row: list) -> None:
        ws = self._sheet(tab, header, create=True)
        ws.append(list(row))

    # -- non-destructive upsert -------------------------------------------- #
    def _find_or_write_header(self, ws, header, key_cols) -> tuple[list[str], int]:
        """Locate the header row (tolerating banner rows above it, like the
        importer) or write one. Returns (normalized labels, 1-based row)."""
        maxrow, maxcol = ws.max_row, ws.max_column
        for r in range(1, min(maxrow, 15) + 1):
            labels = [_norm(ws.cell(row=r, column=c).value) for c in range(1, maxcol + 1)]
            if all(_match_col(labels, c) >= 0 for c in key_cols):
                return labels, r
        empty = all(
            ws.cell(row=r, column=c).value is None
            for r in range(1, maxrow + 1) for c in range(1, maxcol + 1)
        )
        hr = 1 if empty else maxrow + 2  # never overwrite existing content
        for j, label in enumerate(header, start=1):
            ws.cell(row=hr, column=j, value=label)
        return [_norm(l) for l in header], hr

    def upsert_tab(self, tab: str, header: list[str], rows: list[list],
                   key_cols: tuple[str, ...]) -> dict:
        """Update matching rows in place and append new ones, keyed on
        ``key_cols`` — never clears the tab, so operator-added rows and columns
        survive. Falls back to replace when no key is defined."""
        if not key_cols:
            self.replace_tab(tab, header, rows)
            return {"updated": 0, "appended": len(rows)}

        ws = self.wb[tab] if tab in self.wb.sheetnames else self.wb.create_sheet(title=tab)
        labels, header_row = self._find_or_write_header(ws, header, key_cols)

        # Column (1-based) for each owned label; allocate missing ones at the end
        # so we never overwrite an operator column we don't recognize.
        colpos: dict[str, int] = {}
        for label in header:
            j = _match_col(labels, label)
            if j < 0:
                j = len(labels)
                labels.append(_norm(label))
                ws.cell(row=header_row, column=j + 1, value=label)
            colpos[label] = j + 1

        key_at = [header.index(c) for c in key_cols]
        key_cols_at = [colpos[c] for c in key_cols]
        existing: dict[tuple, int] = {}
        for r in range(header_row + 1, ws.max_row + 1):
            k = tuple(layout._norm_key(ws.cell(row=r, column=ci).value) for ci in key_cols_at)
            if any(k):
                existing.setdefault(k, r)

        updated = appended = 0
        for row in rows:
            k = tuple(layout._norm_key(row[p]) for p in key_at)
            target = existing.get(k)
            if target is None:
                target = ws.max_row + 1
                existing[k] = target
                appended += 1
            else:
                updated += 1
            for label, value in zip(header, row):
                ws.cell(row=target, column=colpos[label], value=value)
        return {"updated": updated, "appended": appended}

    def to_bytes(self) -> bytes:
        buf = io.BytesIO()
        self.wb.save(buf)
        return buf.getvalue()

    def close(self) -> None:
        self.wb.close()


class InMemoryXlsxWriter(_WorkbookWriter):
    """Build a fresh .xlsx entirely in memory — backs the download endpoint."""

    def __init__(self):
        super().__init__(self._new_or_load(None))

    def commit(self) -> None:  # nothing to flush; caller reads to_bytes()
        pass


class XlsxWriter(_WorkbookWriter):
    """Read-modify-write a local .xlsx. Preserves tabs the app doesn't own."""

    def __init__(self, path: str):
        self.path = path
        data = None
        if os.path.exists(path):
            with open(path, "rb") as fh:
                data = fh.read()
        super().__init__(self._new_or_load(data))

    def commit(self) -> None:
        self.wb.save(self.path)


class DriveXlsxWriter(_WorkbookWriter):
    """Read-modify-write an .xlsx that lives on Google Drive."""

    def __init__(self, file_id: str, token: str):
        self.file_id = file_id
        self.token = token
        data = source.load_bytes_from_drive(file_id, token)
        super().__init__(self._new_or_load(data))

    def commit(self) -> None:
        source.upload_xlsx_to_drive(self.file_id, self.to_bytes(), self.token)


# --------------------------------------------------------------------------- #
# Native Google Sheets writer (Sheets API v4)
# --------------------------------------------------------------------------- #
_SHEETS_API = "https://sheets.googleapis.com/v4/spreadsheets/{id}"


class GoogleSheetsWriter:
    """Write into a live Google Sheet cell-by-cell via the values API.

    Unlike the .xlsx writers there's no local workbook: each call is an HTTP
    request, so ``commit`` is a no-op. Missing tabs are created on demand.
    """

    def __init__(self, spreadsheet_id: str, token: str, *, timeout: float = 30.0):
        import httpx  # local import keeps httpx optional for pure-xlsx use

        self.id = spreadsheet_id
        self.base = _SHEETS_API.format(id=spreadsheet_id)
        self.client = httpx.Client(
            timeout=timeout,
            headers={"Authorization": f"Bearer {token}",
                     "Content-Type": "application/json"},
        )
        self._titles: set[str] | None = None

    # -- tab bookkeeping ---------------------------------------------------- #
    def _tab_titles(self) -> set[str]:
        if self._titles is None:
            resp = self.client.get(self.base, params={"fields": "sheets.properties.title"})
            resp.raise_for_status()
            self._titles = {
                s["properties"]["title"] for s in resp.json().get("sheets", [])
            }
        return self._titles

    def _ensure_tab(self, tab: str) -> None:
        if tab in self._tab_titles():
            return
        resp = self.client.post(
            f"{self.base}:batchUpdate",
            json={"requests": [{"addSheet": {"properties": {"title": tab}}}]},
        )
        resp.raise_for_status()
        self._tab_titles().add(tab)

    @staticmethod
    def a1_range(tab: str, cell: str | None = None) -> str:
        """URL-safe A1 range for a tab. Sheet names with spaces or special
        chars (all our tabs — "Grow Cycle Log", "Buyers & Pricing") must be
        single-quoted in A1 notation and percent-encoded in the request path;
        without this the Sheets API rejects the range with a 400."""
        name = "'" + tab.replace("'", "''") + "'"
        rng = f"{name}!{cell}" if cell else name
        return urllib.parse.quote(rng, safe="")

    @staticmethod
    def _values(header: list[str], rows: list[list]) -> list[list[str]]:
        out = [list(header)]
        out.extend([layout.cell_to_str(c) for c in row] for row in rows)
        return out

    # -- SheetWriter interface --------------------------------------------- #
    def replace_tab(self, tab: str, header: list[str], rows: list[list]) -> None:
        self._ensure_tab(tab)
        # Clear the whole tab, then write header + rows from A1.
        self.client.post(
            f"{self.base}/values/{self.a1_range(tab)}:clear"
        ).raise_for_status()
        body = {"values": self._values(header, rows)}
        resp = self.client.put(
            f"{self.base}/values/{self.a1_range(tab, 'A1')}",
            params={"valueInputOption": "USER_ENTERED"},
            json=body,
        )
        resp.raise_for_status()

    def append_row(self, tab: str, header: list[str], row: list) -> None:
        first_time = tab not in self._tab_titles()
        self._ensure_tab(tab)
        values = [] if not first_time else [list(header)]
        values.append([layout.cell_to_str(c) for c in row])
        resp = self.client.post(
            f"{self.base}/values/{self.a1_range(tab, 'A1')}:append",
            params={"valueInputOption": "USER_ENTERED", "insertDataOption": "INSERT_ROWS"},
            json={"values": values},
        )
        resp.raise_for_status()

    def _read_values(self, tab: str) -> list[list[str]]:
        resp = self.client.get(
            f"{self.base}/values/{self.a1_range(tab)}",
            params={"majorDimension": "ROWS"},
        )
        resp.raise_for_status()
        return resp.json().get("values", [])

    def upsert_tab(self, tab: str, header: list[str], rows: list[list],
                   key_cols: tuple[str, ...]) -> dict:
        """Keyed upsert against a live Sheet: update matching rows in place,
        append new ones, and leave unmanaged rows/columns intact. Owned columns
        are written as one contiguous span per row so operator columns to the
        right are preserved (see append fallback when there's no header yet)."""
        if not key_cols:
            self.replace_tab(tab, header, rows)
            return {"updated": 0, "appended": len(rows)}

        self._ensure_tab(tab)
        values = self._read_values(tab)

        header_idx = None
        for i, rowvals in enumerate(values[:15]):
            labels = [_norm(v) for v in rowvals]
            if all(_match_col(labels, c) >= 0 for c in key_cols):
                header_idx = i
                break

        # No header found -> fresh tab: append header + every row.
        if header_idx is None:
            body = [list(header)] + [[layout.cell_to_str(c) for c in r] for r in rows]
            self.client.post(
                f"{self.base}/values/{self.a1_range(tab, 'A1')}:append",
                params={"valueInputOption": "USER_ENTERED", "insertDataOption": "INSERT_ROWS"},
                json={"values": body},
            ).raise_for_status()
            return {"updated": 0, "appended": len(rows)}

        labels = [_norm(v) for v in values[header_idx]]
        colpos: dict[str, int] = {}
        header_updates: list[dict] = []
        for label in header:
            j = _match_col(labels, label)
            if j < 0:
                j = len(labels)
                labels.append(_norm(label))
                header_updates.append({
                    "range": self._a1_cell(tab, header_idx, j),
                    "values": [[label]],
                })
            colpos[label] = j

        lo, hi = min(colpos.values()), max(colpos.values())
        key_at = [header.index(c) for c in key_cols]
        key_cols_pos = [colpos[c] for c in key_cols]

        existing: dict[tuple, int] = {}
        for ri in range(header_idx + 1, len(values)):
            rowvals = values[ri]
            k = tuple(
                layout._norm_key(rowvals[p] if p < len(rowvals) else "")
                for p in key_cols_pos
            )
            if any(k):
                existing.setdefault(k, ri)

        data = list(header_updates)
        appends: list[list[str]] = []
        next_ri = len(values)
        updated = appended = 0
        for row in rows:
            k = tuple(layout._norm_key(row[p]) for p in key_at)
            span = [""] * (hi - lo + 1)
            for label, value in zip(header, row):
                span[colpos[label] - lo] = layout.cell_to_str(value)
            ri = existing.get(k)
            if ri is None:
                existing[k] = next_ri
                next_ri += 1
                appended += 1
                # Pad so the span sits at the right columns when appended.
                appends.append([""] * lo + span)
            else:
                updated += 1
            row_at = existing[k]
            if ri is not None:
                data.append({
                    "range": self._a1_span(tab, row_at, lo, hi),
                    "values": [span],
                })
        if data:
            self.client.post(
                f"{self.base}/values:batchUpdate",
                json={"valueInputOption": "USER_ENTERED", "data": data},
            ).raise_for_status()
        if appends:
            self.client.post(
                f"{self.base}/values/{self.a1_range(tab, 'A1')}:append",
                params={"valueInputOption": "USER_ENTERED", "insertDataOption": "INSERT_ROWS"},
                json={"values": appends},
            ).raise_for_status()
        return {"updated": updated, "appended": appended}

    def _a1_cell(self, tab: str, row_idx0: int, col_idx0: int) -> str:
        cell = f"{_col_letter(col_idx0)}{row_idx0 + 1}"
        return self.a1_range(tab, cell)

    def _a1_span(self, tab: str, row_idx0: int, lo: int, hi: int) -> str:
        cell = f"{_col_letter(lo)}{row_idx0 + 1}:{_col_letter(hi)}{row_idx0 + 1}"
        return self.a1_range(tab, cell)

    def commit(self) -> None:
        pass

    def close(self) -> None:
        self.client.close()


# --------------------------------------------------------------------------- #
# Environment-driven backend selection
# --------------------------------------------------------------------------- #
# Ordered write backends: (env var, kind, needs a Google token). describe_target
# and resolve_writer both consume this so their precedence can't drift.
_WRITE_BACKENDS = (
    ("MASTER_SHEET_GOOGLE_ID", "google_sheet", True),
    ("MASTER_SHEET_FILE_ID", "drive_xlsx", True),
    ("MASTER_SHEET_PATH", "local_xlsx", False),
)


def describe_target() -> dict:
    """Which write backend the environment resolves to, without building it or
    minting a token. Powers /api/sync/status so the UI can tell the operator
    whether write-back is wired up and to where.
    """
    for var, kind, needs_token in _WRITE_BACKENDS:
        ref = os.environ.get(var)
        if ref:
            # credentials_available() only checks env presence — no network.
            writable = (not needs_token) or source.credentials_available()
            return {"configured": True, "kind": kind, "ref": ref, "writable": writable}
    return {"configured": False, "kind": None, "ref": None, "writable": False}


def resolve_writer(*, path: str | None = None, token: str | None = None) -> SheetWriter:
    """Build the configured write backend. Priority mirrors the read source:
    a native Google Sheet id, then an .xlsx on Drive, then a local .xlsx.

    Raises RuntimeError with an actionable message when nothing is configured.
    """
    def _token(var: str) -> str:
        tok = source.resolve_write_token(token=token)
        if not tok:
            raise RuntimeError(
                f"{var} is set but no Google credentials are available. "
                "Set GOOGLE_SERVICE_ACCOUNT_JSON or GOOGLE_OAUTH_TOKEN."
            )
        return tok

    google_id = os.environ.get("MASTER_SHEET_GOOGLE_ID")
    if google_id:
        return GoogleSheetsWriter(google_id, _token("MASTER_SHEET_GOOGLE_ID"))

    file_id = os.environ.get("MASTER_SHEET_FILE_ID")
    if file_id:
        return DriveXlsxWriter(file_id, _token("MASTER_SHEET_FILE_ID"))

    local = path or os.environ.get("MASTER_SHEET_PATH")
    if local:
        return XlsxWriter(local)

    raise RuntimeError(
        "No write target configured. Set MASTER_SHEET_GOOGLE_ID (a Google "
        "Sheet), MASTER_SHEET_FILE_ID (an .xlsx on Drive), or MASTER_SHEET_PATH "
        "(a local .xlsx)."
    )
