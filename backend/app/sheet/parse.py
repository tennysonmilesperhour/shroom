"""Parse the Master Cultivation Reference workbook into normalized records.

Each tab becomes a list of plain dataclasses with stable, target-agnostic field
names; the sinks (SQLite / Supabase) decide how to persist them. Parsing is
defensive: headers are located by scanning for known tokens (so inserting a
column or a banner row doesn't break the import), banner/total rows are skipped,
and an unparseable cell degrades to empty rather than raising.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date

from openpyxl.workbook import Workbook

from . import util


# --------------------------------------------------------------------------- #
# Records
# --------------------------------------------------------------------------- #
@dataclass
class Strain:
    name: str
    mushroom_type: str = "psychedelic"
    species: str = ""
    vendor: str = ""
    potency: str = ""
    ease_rating: int | None = None
    grow_again: bool | None = None
    library_status: str = ""
    priority: int | None = None
    syringes_on_hand: float | None = None
    acquired_on: date | None = None
    container_id: str = ""
    notes: str = ""


@dataclass
class Vendor:
    name: str
    category: str = "supplies"
    products: str = ""
    url: str = ""
    rating: int | None = None
    contact_priority: str = ""
    notes: str = ""


@dataclass
class Equipment:
    name: str
    spec_notes: str = ""
    status: str = "active"
    last_checked: str = ""


@dataclass
class Customer:
    name: str
    channel: str = "wholesale"
    status: str = "lead"
    role: str = ""
    price_tier: str = ""
    volume_est: str = ""
    last_contact: date | None = None
    notes: str = ""


@dataclass
class PriceTier:
    tier: str
    product_class: str = "medicinal"
    min_per_gram: float | None = None
    max_per_gram: float | None = None
    notes: str = ""


@dataclass
class Jar:
    jar_id: str
    strain: str = ""
    flush_number: int | None = None
    dry_weight_g: float = 0.0
    used_g: float = 0.0
    notes: str = ""


@dataclass
class SourcedGood:
    strain: str
    on_hand_g: float = 0.0
    used_g: float = 0.0
    incoming: str = ""
    last_updated: date | None = None
    notes: str = ""


@dataclass
class Sale:
    # Natural key for idempotent upserts: (sale_date, buyer, strains, amount).
    sale_date: date | None
    buyer: str
    strains: str
    grams: float | None
    amount: float | None
    tier: str = ""
    source_notes: str = ""
    payment: str = ""


@dataclass
class Protocol:
    name: str
    category: str = "sop"
    steps: list[str] = field(default_factory=list)


@dataclass
class Guide:
    guide_type: str          # 'contamination' | 'symptom'
    label: str
    appearance: str = ""
    cause: str = ""
    action: str = ""


@dataclass
class Incident:
    # Natural key for idempotent upserts: (log_date, issue).
    log_date: date | None
    issue: str
    root_cause: str = ""
    resolution: str = ""


@dataclass
class Batch:
    lot_code: str
    strain: str
    container_id: str = ""
    flush_number: int | None = None
    stage: str = "inoculation"
    inoculated_on: date | None = None
    transferred_on: date | None = None
    first_pins_on: date | None = None
    contamination_flag: bool = False
    issues: str = ""
    notes: str = ""


@dataclass
class Harvest:
    lot_code: str           # links to the batch synthesized from tub+flush
    strain: str
    flush_number: int
    harvested_on: date | None
    fresh_g: float = 0.0
    dry_g: float = 0.0
    notes: str = ""


@dataclass
class ParsedWorkbook:
    strains: list[Strain] = field(default_factory=list)
    vendors: list[Vendor] = field(default_factory=list)
    equipment: list[Equipment] = field(default_factory=list)
    customers: list[Customer] = field(default_factory=list)
    price_tiers: list[PriceTier] = field(default_factory=list)
    jars: list[Jar] = field(default_factory=list)
    sourced_goods: list[SourcedGood] = field(default_factory=list)
    sales: list[Sale] = field(default_factory=list)
    protocols: list[Protocol] = field(default_factory=list)
    guides: list[Guide] = field(default_factory=list)
    incidents: list[Incident] = field(default_factory=list)
    batches: list[Batch] = field(default_factory=list)
    harvests: list[Harvest] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Low-level sheet helpers
# --------------------------------------------------------------------------- #
def _get_sheet(wb: Workbook, *candidates: str):
    """Find a worksheet by fuzzy (case/space-insensitive substring) name."""
    norm = {re.sub(r"\s+", "", s.lower()): s for s in wb.sheetnames}
    for cand in candidates:
        key = re.sub(r"\s+", "", cand.lower())
        for k, original in norm.items():
            if key in k or k in key:
                return wb[original]
    return None


def _matrix(ws) -> list[list]:
    if ws is None:
        return []
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _row_text(row: list) -> str:
    return " | ".join(util.clean(c) for c in row).lower()


def _find_header(matrix: list[list], tokens: list[str], start: int = 0) -> int:
    """Index of the first row (>= start) whose cells contain every token."""
    want = [t.lower() for t in tokens]
    for i in range(start, len(matrix)):
        text = _row_text(matrix[i])
        if all(t in text for t in want):
            return i
    return -1


def _col(headers: list, *aliases: str) -> int:
    """Column index of the first header matching any alias (substring)."""
    cleaned = [util.clean(h).lower() for h in headers]
    for alias in aliases:
        a = alias.lower()
        for i, h in enumerate(cleaned):
            if a in h:
                return i
    return -1


def _at(row: list, idx: int):
    return row[idx] if 0 <= idx < len(row) else None


def _is_blank_row(row: list) -> bool:
    return all(util.is_blank(c) for c in row)


def _non_blank_count(row: list) -> int:
    return sum(1 for c in row if not util.is_blank(c))


# --------------------------------------------------------------------------- #
# Per-tab parsers
# --------------------------------------------------------------------------- #
# Only a *container* parenthetical is noise ("Golden Teacher (bag)"); a name
# alias like "BV (Bluey Vuittons)" is meaningful and must be preserved.
_CONTAINER_PAREN = re.compile(
    r"\s*\([^)]*\b(?:bag|grain|aio|tub|monotub|jar)\b[^)]*\)\s*$", re.I)

# Leading list decoration the sheet uses for bulleted strain rows: "• PE6",
# "- Enigma", "* Fiji". If left on, "• PE6" and "PE6" are distinct names and
# the importer inserts a second row instead of upserting onto the first — the
# source of the duplicate "• Name" strains. Stripped so both collapse to one.
_LEADING_DECOR = re.compile(r"^[\s•▪◦‣·♦●○*\-–—]+")

# Rows that are order/shipment notes parked in a strain column ("NEW SPORES —
# order #6849 shipped", "tracking 1Z…"), not actual strains. Kept out of the
# library so they don't masquerade as cultures.
_ORDER_NOTE = re.compile(
    r"(?i)\border\s*#|#\d{3,}|\btracking\b|\bshipped\b|\bshipment\b|\beta\b|"
    r"\bnew\s+spores\b")


def _strip_name(name: str) -> str:
    """Canonical strain name: drop a trailing container parenthetical and any
    leading bullet/list decoration, then collapse internal whitespace."""
    cleaned = _CONTAINER_PAREN.sub("", util.clean(name))
    cleaned = _LEADING_DECOR.sub("", cleaned)
    return re.sub(r"\s{2,}", " ", cleaned).strip()


def _is_order_note(name: str) -> bool:
    """True for cells that are order/shipment notes rather than strain names."""
    return bool(_ORDER_NOTE.search(name))


def parse_strain_library(wb: Workbook) -> list[Strain]:
    ws = _get_sheet(wb, "Strain Library")
    matrix = _matrix(ws)
    h = _find_header(matrix, ["strain", "status", "potency", "grow again"])
    if h < 0:
        return []
    headers = matrix[h]
    c_name = _col(headers, "strain")
    c_status = _col(headers, "status")
    c_vendor = _col(headers, "vendor")
    c_inoc = _col(headers, "inoculated")
    c_potency = _col(headers, "potency")
    c_ease = _col(headers, "ease")
    c_grow = _col(headers, "grow again")
    c_tub = _col(headers, "tub", "bag id")
    c_notes = _col(headers, "notes")

    out: list[Strain] = []
    for row in matrix[h + 1:]:
        if _is_blank_row(row):
            continue
        name_raw = util.clean(_at(row, c_name))
        # Section banners ("MEDICINAL — COLONIZING") have a name but no status.
        if not name_raw or util.is_blank(_at(row, c_status)):
            continue
        # Order/shipment notes sometimes sit in the strain column — not cultures.
        if _is_order_note(name_raw):
            continue
        notes = util.clean(_at(row, c_notes))
        out.append(Strain(
            name=_strip_name(name_raw),
            mushroom_type="psychedelic",
            species=util.species_from_notes(notes) or "Psilocybe cubensis",
            vendor=util.clean(_at(row, c_vendor)),
            potency=util.clean(_at(row, c_potency)),
            ease_rating=util.parse_rating(_at(row, c_ease)),
            grow_again=util.parse_bool(_at(row, c_grow)),
            library_status=util.library_status(_at(row, c_status)),
            acquired_on=util.parse_date(_at(row, c_inoc)),
            container_id=util.clean(_at(row, c_tub)),
            notes=notes,
        ))
    return out


def parse_fridge_and_incoming(wb: Workbook) -> list[Strain]:
    ws = _get_sheet(wb, "Fridge & Incoming", "Fridge")
    matrix = _matrix(ws)
    out: list[Strain] = []

    # Sub-table 1: fridge spore inventory (psychedelic).
    h = _find_header(matrix, ["strain", "location", "qty", "priority"])
    if h >= 0:
        headers = matrix[h]
        c_name = _col(headers, "strain")
        c_status = _col(headers, "status")
        c_qty = _col(headers, "qty")
        c_potency = _col(headers, "potency")
        c_ease = _col(headers, "ease")
        c_prio = _col(headers, "priority")
        c_vendor = _col(headers, "vendor")
        c_acq = _col(headers, "acquired")
        c_notes = _col(headers, "notes")
        for row in matrix[h + 1:]:
            if _is_blank_row(row):
                continue
            text = _row_text(row)
            # The next sub-table ("FUNCTIONAL ... INCOMING") starts a new header.
            if "incoming" in text and "north spore" in text:
                break
            name = util.clean(_at(row, c_name))
            if not name or util.is_blank(_at(row, c_status)):
                continue
            if _is_order_note(name):
                continue
            out.append(Strain(
                name=_strip_name(name),
                mushroom_type="psychedelic",
                vendor=util.clean(_at(row, c_vendor)),
                potency=util.clean(_at(row, c_potency)),
                ease_rating=util.parse_rating(_at(row, c_ease)),
                library_status=util.library_status(_at(row, c_status)) or "fridge",
                priority=util.parse_stars(_at(row, c_prio)),
                syringes_on_hand=util.first_number(_at(row, c_qty)),
                acquired_on=util.parse_date(_at(row, c_acq)),
                notes=util.clean(_at(row, c_notes)),
            ))

    # Sub-table 2: functional / specialty incoming.
    h2 = _find_header(matrix, ["strain", "format", "use", "market"])
    if h2 >= 0:
        headers = matrix[h2]
        c_name = _col(headers, "strain")
        c_status = _col(headers, "status")
        c_vendor = _col(headers, "vendor")
        c_ease = _col(headers, "ease")
        c_notes = _col(headers, "notes")
        for row in matrix[h2 + 1:]:
            if _is_blank_row(row):
                continue
            name = util.clean(_at(row, c_name))
            text = _row_text(row)
            if not name:
                continue
            if "supplies" in text and "substrate" in text:
                break
            if util.is_blank(_at(row, c_status)):
                continue
            if _is_order_note(name):
                continue
            out.append(Strain(
                name=_strip_name(name),
                mushroom_type="functional",
                vendor=util.clean(_at(row, c_vendor)) or "North Spore",
                ease_rating=util.parse_rating(_at(row, c_ease)),
                library_status=util.library_status(_at(row, c_status)) or "en_route",
                notes=util.clean(_at(row, c_notes)),
            ))
    return out


def parse_vendors(wb: Workbook) -> list[Vendor]:
    ws = _get_sheet(wb, "Vendors")
    matrix = _matrix(ws)
    out: list[Vendor] = []

    h = _find_header(matrix, ["vendor", "products", "rating"])
    sourcing_start = _find_header(matrix, ["vendor", "region", "contact priority"])
    end_supply = sourcing_start if sourcing_start > 0 else len(matrix)
    if h >= 0:
        headers = matrix[h]
        c_name = _col(headers, "vendor")
        c_prod = _col(headers, "products")
        c_url = _col(headers, "url")
        c_rating = _col(headers, "rating")
        c_notes = _col(headers, "notes")
        for row in matrix[h + 1:end_supply]:
            name = util.clean(_at(row, c_name))
            if not name or name.upper() == name and _non_blank_count(row) <= 1:
                continue
            products = util.clean(_at(row, c_prod))
            cat = "spores" if re.search(r"spore|syringe|\bLC\b|spawn|print", products, re.I) else "supplies"
            out.append(Vendor(
                name=name, category=cat, products=products,
                url=util.clean(_at(row, c_url)),
                rating=util.parse_stars(_at(row, c_rating)),
                notes=util.clean(_at(row, c_notes)),
            ))

    if sourcing_start >= 0:
        headers = matrix[sourcing_start]
        c_name = _col(headers, "vendor")
        c_region = _col(headers, "region")
        c_url = _col(headers, "url")
        c_prio = _col(headers, "contact priority")
        c_notes = _col(headers, "notes")
        for row in matrix[sourcing_start + 1:]:
            name = util.clean(_at(row, c_name))
            if not name or "note" in name.lower():
                continue
            out.append(Vendor(
                name=name, category="sourcing",
                products=util.clean(_at(row, c_region)),
                url=util.clean(_at(row, c_url)),
                contact_priority=util.clean(_at(row, c_prio)),
                notes=util.clean(_at(row, c_notes)),
            ))
    return out


def parse_equipment(wb: Workbook) -> list[Equipment]:
    ws = _get_sheet(wb, "Environment")
    matrix = _matrix(ws)
    h = _find_header(matrix, ["equipment", "status", "last checked"])
    if h < 0:
        return []
    headers = matrix[h]
    c_name = _col(headers, "equipment")
    c_spec = _col(headers, "spec", "notes")
    c_status = _col(headers, "status")
    c_checked = _col(headers, "last checked")
    out: list[Equipment] = []
    for row in matrix[h + 1:]:
        text = _row_text(row)
        if "environmental targets" in text or "parameter" in text and "target" in text:
            break
        name = util.clean(_at(row, c_name))
        if not name or util.is_blank(_at(row, c_spec)):
            continue
        out.append(Equipment(
            name=name,
            spec_notes=util.clean(_at(row, c_spec)),
            status=util.clean(_at(row, c_status)) or "active",
            last_checked=util.clean(_at(row, c_checked)),
        ))
    return out


def parse_buyers(wb: Workbook) -> list[Customer]:
    ws = _get_sheet(wb, "Buyers & Pricing", "Buyers")
    matrix = _matrix(ws)
    h = _find_header(matrix, ["name", "tier", "role", "status"])
    if h < 0:
        return []
    headers = matrix[h]
    c_name = _col(headers, "name")
    c_tier = _col(headers, "tier")
    c_role = _col(headers, "role")
    c_vol = _col(headers, "volume")
    c_last = _col(headers, "last contact")
    c_status = _col(headers, "status")
    c_notes = _col(headers, "notes")
    out: list[Customer] = []
    for row in matrix[h + 1:]:
        name = util.clean(_at(row, c_name))
        if not name or name.upper() == name and _non_blank_count(row) <= 1:
            continue
        tier = util.clean(_at(row, c_tier))
        out.append(Customer(
            name=name,
            channel=util.channel_for_tier(tier),
            status=util.customer_status(_at(row, c_status)),
            role=util.clean(_at(row, c_role)),
            price_tier=tier,
            volume_est=util.clean(_at(row, c_vol)),
            last_contact=util.parse_date(_at(row, c_last)),
            notes=util.clean(_at(row, c_notes)),
        ))
    return out


def parse_jar_inventory(wb: Workbook) -> tuple[list[Jar], list[PriceTier]]:
    ws = _get_sheet(wb, "Jar Inventory", "Jar")
    matrix = _matrix(ws)
    jars: list[Jar] = []
    h = _find_header(matrix, ["jar id", "strain", "dry", "remaining"])
    if h >= 0:
        headers = matrix[h]
        c_jar = _col(headers, "jar id")
        c_strain = _col(headers, "strain")
        c_flush = _col(headers, "flush")
        c_dry = _col(headers, "dry")
        c_used = _col(headers, "used")
        c_notes = _col(headers, "notes")
        for row in matrix[h + 1:]:
            jar_id = util.clean(_at(row, c_jar))
            if not jar_id or jar_id.upper().startswith("TOTAL"):
                if jar_id.upper().startswith("TOTAL"):
                    break
                continue
            jars.append(Jar(
                jar_id=jar_id,
                strain=_strip_name(_at(row, c_strain)),
                flush_number=util.parse_int(_at(row, c_flush)),
                dry_weight_g=util.first_number(_at(row, c_dry)) or 0.0,
                used_g=util.first_number(_at(row, c_used)) or 0.0,
                notes=util.clean(_at(row, c_notes)),
            ))

    tiers: list[PriceTier] = []
    ph = _find_header(matrix, ["tier", "volume", "price"])
    if ph >= 0:
        headers = matrix[ph]
        c_tier = _col(headers, "tier")
        c_vol = _col(headers, "volume")
        c_price = _col(headers, "price")
        c_notes = _col(headers, "notes")
        for row in matrix[ph + 1:]:
            tier = util.clean(_at(row, c_tier))
            if not tier:
                continue
            lo, hi = util.money_range(_at(row, c_price))
            note = " · ".join(p for p in [util.clean(_at(row, c_vol)), util.clean(_at(row, c_notes))] if p)
            tiers.append(PriceTier(
                tier=tier.lower(), product_class="medicinal",
                min_per_gram=lo, max_per_gram=hi, notes=note,
            ))
    return jars, tiers


def parse_sourced_goods(wb: Workbook) -> list[SourcedGood]:
    ws = _get_sheet(wb, "Sourced Finished Goods", "Sourced")
    matrix = _matrix(ws)
    h = _find_header(matrix, ["strain", "on-hand", "remaining"])
    if h < 0:
        return []
    headers = matrix[h]
    c_strain = _col(headers, "strain")
    c_on = _col(headers, "on-hand", "on hand")
    c_used = _col(headers, "used")
    c_incoming = _col(headers, "incoming")
    c_updated = _col(headers, "last updated")
    c_notes = _col(headers, "notes")
    out: list[SourcedGood] = []
    for row in matrix[h + 1:]:
        strain = util.clean(_at(row, c_strain))
        if not strain:
            continue
        out.append(SourcedGood(
            strain=_strip_name(strain),
            on_hand_g=util.first_number(_at(row, c_on)) or 0.0,
            used_g=util.first_number(_at(row, c_used)) or 0.0,
            incoming=util.clean(_at(row, c_incoming)),
            last_updated=util.parse_date(_at(row, c_updated)),
            notes=util.clean(_at(row, c_notes)),
        ))
    return out


_SALE_SKIP = ("total", "collected", "outstanding")


def parse_sales(wb: Workbook) -> list[Sale]:
    ws = _get_sheet(wb, "Sales Log", "Sales")
    matrix = _matrix(ws)
    h = _find_header(matrix, ["date", "buyer", "grams", "amount"])
    if h < 0:
        return []
    headers = matrix[h]
    c_date = _col(headers, "date")
    c_buyer = _col(headers, "buyer")
    c_strain = _col(headers, "strain")
    c_grams = _col(headers, "grams")
    c_amt = _col(headers, "amount")
    c_tier = _col(headers, "tier")
    c_src = _col(headers, "source", "notes")
    c_pay = _col(headers, "payment")
    out: list[Sale] = []
    for row in matrix[h + 1:]:
        buyer = util.clean(_at(row, c_buyer))
        first = util.clean(_at(row, c_date))
        if first.lower().rstrip(":") in _SALE_SKIP or buyer.lower().rstrip(":") in _SALE_SKIP:
            continue
        if not buyer:
            continue
        out.append(Sale(
            sale_date=util.parse_date(_at(row, c_date)),
            buyer=buyer,
            strains=util.clean(_at(row, c_strain)),
            grams=util.grams(_at(row, c_grams)),
            amount=util.first_number(_at(row, c_amt)),
            tier=util.clean(_at(row, c_tier)),
            source_notes=util.clean(_at(row, c_src)),
            payment=util.clean(_at(row, c_pay)),
        ))
    return out


_CHECKBOX = re.compile(r"^\s*\[\s*[xX]?\s*\]")


def parse_protocols(wb: Workbook) -> list[Protocol]:
    ws = _get_sheet(wb, "Protocols")
    matrix = _matrix(ws)
    out: list[Protocol] = []
    current: Protocol | None = None
    for row in matrix:
        if _is_blank_row(row):
            continue
        first = util.clean(_at(row, 0))
        rest = " ".join(util.clean(c) for c in row[1:] if not util.is_blank(c))
        if _CHECKBOX.match(first) or (not first and rest):
            # A step line: text is whatever follows the checkbox cell.
            step = rest or _CHECKBOX.sub("", first).strip()
            if current is not None and step:
                current.steps.append(step)
        elif first:
            # A new protocol heading.
            current = Protocol(name=first.title() if first.isupper() else first)
            out.append(current)
    return [p for p in out if p.steps]


def parse_troubleshooting(wb: Workbook) -> tuple[list[Guide], list[Incident]]:
    ws = _get_sheet(wb, "Troubleshooting")
    matrix = _matrix(ws)
    guides: list[Guide] = []
    incidents: list[Incident] = []

    ch = _find_header(matrix, ["type", "appearance", "cause", "action"])
    sh = _find_header(matrix, ["symptom", "cause", "fix"])
    ih = _find_header(matrix, ["date", "issue", "root cause", "resolution"])

    bounds = sorted(x for x in [ch, sh, ih, len(matrix)] if x >= 0)

    def next_bound(after: int) -> int:
        for b in bounds:
            if b > after:
                return b
        return len(matrix)

    if ch >= 0:
        headers = matrix[ch]
        c_type = _col(headers, "type")
        c_app = _col(headers, "appearance")
        c_cause = _col(headers, "cause")
        c_act = _col(headers, "action")
        for row in matrix[ch + 1:next_bound(ch)]:
            label = util.clean(_at(row, c_type))
            if not label or util.is_blank(_at(row, c_act)):
                continue
            guides.append(Guide("contamination", label,
                                appearance=util.clean(_at(row, c_app)),
                                cause=util.clean(_at(row, c_cause)),
                                action=util.clean(_at(row, c_act))))

    if sh >= 0:
        headers = matrix[sh]
        c_sym = _col(headers, "symptom")
        c_cause = _col(headers, "likely cause", "cause")
        c_fix = _col(headers, "fix")
        for row in matrix[sh + 1:next_bound(sh)]:
            label = util.clean(_at(row, c_sym))
            if not label or util.is_blank(_at(row, c_fix)):
                continue
            guides.append(Guide("symptom", label,
                                cause=util.clean(_at(row, c_cause)),
                                action=util.clean(_at(row, c_fix))))

    if ih >= 0:
        headers = matrix[ih]
        c_date = _col(headers, "date")
        c_issue = _col(headers, "issue")
        c_root = _col(headers, "root cause")
        c_res = _col(headers, "resolution")
        for row in matrix[ih + 1:next_bound(ih)]:
            issue = util.clean(_at(row, c_issue))
            if not issue:
                continue
            incidents.append(Incident(
                log_date=util.parse_date(_at(row, c_date)),
                issue=issue,
                root_cause=util.clean(_at(row, c_root)),
                resolution=util.clean(_at(row, c_res)),
            ))
    return guides, incidents


def _lot_code(tub: str, flush: str | int | None) -> str:
    tub = util.clean(tub) or "UNK"
    f = util.parse_int(flush)
    return f"{tub}-F{f}" if f else tub


def parse_grow_cycle(wb: Workbook) -> list[Batch]:
    ws = _get_sheet(wb, "Grow Cycle Log", "Grow Cycle", "Cycle Log")
    matrix = _matrix(ws)
    h = _find_header(matrix, ["strain", "tub", "flush", "inoculated"])
    if h < 0:
        return []
    headers = matrix[h]
    c_strain = _col(headers, "strain")
    c_tub = _col(headers, "tub")
    c_flush = _col(headers, "flush")
    c_inoc = _col(headers, "inoculated")
    c_trans = _col(headers, "transferred")
    c_pins = _col(headers, "first pins")
    c_harv = _col(headers, "harvest date")
    c_contam = _col(headers, "contam")
    c_issues = _col(headers, "issues")
    c_notes = _col(headers, "notes")
    out: list[Batch] = []
    for row in matrix[h + 1:]:
        strain = util.clean(_at(row, c_strain))
        tub = util.clean(_at(row, c_tub))
        if not strain or not tub or strain.upper().startswith("TOTAL"):
            continue
        harvested = util.parse_date(_at(row, c_harv))
        pins = util.parse_date(_at(row, c_pins))
        trans = util.parse_date(_at(row, c_trans))
        if harvested:
            stage = "harvesting"
        elif pins:
            stage = "fruiting"
        elif trans:
            stage = "colonization"
        else:
            stage = "inoculation"
        contam = util.clean(_at(row, c_contam)).lower()
        out.append(Batch(
            lot_code=_lot_code(tub, _at(row, c_flush)),
            strain=_strip_name(strain),
            container_id=tub,
            flush_number=util.parse_int(_at(row, c_flush)),
            stage=stage,
            inoculated_on=util.parse_date(_at(row, c_inoc)),
            transferred_on=trans,
            first_pins_on=pins,
            contamination_flag=bool(contam) and contam not in ("none", "no"),
            issues=util.clean(_at(row, c_issues)),
            notes=util.clean(_at(row, c_notes)),
        ))
    return out


def parse_harvests(wb: Workbook) -> list[Harvest]:
    ws = _get_sheet(wb, "Harvest Tracker", "Harvest")
    matrix = _matrix(ws)
    h = _find_header(matrix, ["strain", "tub", "flush", "fresh"])
    if h < 0:
        return []
    headers = matrix[h]
    c_strain = _col(headers, "strain")
    c_tub = _col(headers, "tub")
    c_flush = _col(headers, "flush")
    c_date = _col(headers, "harvest date")
    c_fresh = _col(headers, "fresh")
    c_dry = _col(headers, "dry (g)", "dry")
    c_notes = _col(headers, "notes")
    out: list[Harvest] = []
    for row in matrix[h + 1:]:
        strain = util.clean(_at(row, c_strain))
        tub = util.clean(_at(row, c_tub))
        if not strain or strain.upper().startswith("TOTAL") or not tub:
            continue
        flush = util.parse_int(_at(row, c_flush)) or 1
        out.append(Harvest(
            lot_code=_lot_code(tub, flush),
            strain=_strip_name(strain),
            flush_number=flush,
            harvested_on=util.parse_date(_at(row, c_date)),
            fresh_g=util.first_number(_at(row, c_fresh)) or 0.0,
            dry_g=util.first_number(_at(row, c_dry)) or 0.0,
            notes=util.clean(_at(row, c_notes)),
        ))
    return out


# --------------------------------------------------------------------------- #
# Top-level
# --------------------------------------------------------------------------- #
def _merge_strains(*groups: list[Strain]) -> list[Strain]:
    """Combine strain rows from multiple tabs, keyed by canonical name. Later
    non-empty fields fill gaps without clobbering an earlier richer value."""
    by_name: dict[str, Strain] = {}
    order: list[str] = []
    for group in groups:
        for s in group:
            key = s.name.lower()
            if key not in by_name:
                by_name[key] = s
                order.append(key)
                continue
            existing = by_name[key]
            for f in s.__dataclass_fields__:
                new = getattr(s, f)
                old = getattr(existing, f)
                if old in (None, "", 0.0) and new not in (None, "", 0.0):
                    setattr(existing, f, new)
    return [by_name[k] for k in order]


def parse_workbook(wb: Workbook) -> ParsedWorkbook:
    jars, tiers = parse_jar_inventory(wb)
    guides, incidents = parse_troubleshooting(wb)
    strains = _merge_strains(parse_strain_library(wb), parse_fridge_and_incoming(wb))
    return ParsedWorkbook(
        strains=strains,
        vendors=parse_vendors(wb),
        equipment=parse_equipment(wb),
        customers=parse_buyers(wb),
        price_tiers=tiers,
        jars=jars,
        sourced_goods=parse_sourced_goods(wb),
        sales=parse_sales(wb),
        protocols=parse_protocols(wb),
        guides=guides,
        incidents=incidents,
        batches=parse_grow_cycle(wb),
        harvests=parse_harvests(wb),
    )
