"""Tests for the reverse sync direction: App -> Sheet (export/writer/autosync).

The core guarantee is a *round-trip*: data the app writes into a workbook must
come back out through the same parser the importer uses. If that holds, a change
made in the app genuinely lands where the sheet reads it — which is what
two-way sync means.
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from backend.app import models
from backend.app.database import Base
from backend.app.sheet import autosync, export, layout, parse, state, writer


# --------------------------------------------------------------------------- #
# Fixtures — a small but relationship-complete app DB
# --------------------------------------------------------------------------- #
@pytest.fixture()
def session(tmp_path):
    engine = create_engine(f"sqlite:///{tmp_path/'wb.db'}", future=True)
    Base.metadata.create_all(bind=engine)
    db = sessionmaker(bind=engine, future=True)()
    try:
        yield db
    finally:
        db.close()
        engine.dispose()


@pytest.fixture()
def seeded(session):
    sg = models.Strain(
        name="Stargazer", mushroom_type="psychedelic", vendor="Sporeworks",
        potency="high", ease_rating=8, grow_again=True, active=True,
        notes="P. cubensis — reliable",
    )
    pe = models.Strain(
        name="Penis Envy", mushroom_type="psychedelic", vendor="Ravenite",
        potency="very high", ease_rating=5, grow_again=True, active=False,
        notes="slow but potent",
    )
    session.add_all([sg, pe])
    session.flush()

    batch = models.Batch(
        lot_code="T-01-F1", strain_id=sg.id, stage="harvesting",
        inoculated_on=date(2026, 5, 1), colonized_on=date(2026, 5, 12),
        fruiting_on=date(2026, 5, 20), contamination_flag=False, notes="clean run",
    )
    session.add(batch)
    session.flush()

    session.add(models.Harvest(
        batch_id=batch.id, harvested_on=date(2026, 5, 29), flush_number=1,
        weight_kg=0.445, dry_weight_kg=0.0468, notes="first flush",
    ))
    session.add(models.Customer(
        name="Daniel Childs", channel="distributor", notes="net-30",
    ))
    session.commit()
    return session


# --------------------------------------------------------------------------- #
# build_tables — DB projected into the workbook layout
# --------------------------------------------------------------------------- #
def test_build_tables_covers_owned_entities(seeded):
    tables = {t.spec.key: t for t in export.build_tables(seeded)}
    assert set(tables) == {"strains", "batches", "harvests", "customers"}
    # Headers carry every token the parser scans for.
    assert tables["strains"].header[0] == "Strain"
    assert len(tables["strains"].rows) == 2
    assert len(tables["harvests"].rows) == 1


def test_lot_code_splits_into_tub_and_flush():
    assert layout.split_lot("T-01-F2") == ("T-01", 2)
    assert layout.split_lot("T-01") == ("T-01", None)


def test_zero_weight_grams_is_not_dropped():
    # A harvest logged before weighing has weight 0.0 — it must render as 0,
    # not a blank cell, so a real zero round-trips.
    assert layout._grams(0.0) == 0.0
    assert layout._grams(None) is None
    assert layout._grams(0.445) == 445


def test_google_sheets_a1_range_is_quoted_and_encoded():
    # Sheet names with spaces / '&' must be single-quoted in A1 notation and
    # percent-encoded in the request path, or the Sheets API 400s.
    rng = writer.GoogleSheetsWriter.a1_range("Buyers & Pricing", "A1")
    # single-quoted, space -> %20, '&' -> %26, '!' -> %21
    assert rng == "%27Buyers%20%26%20Pricing%27%21A1"
    assert writer.GoogleSheetsWriter.a1_range("Grow Cycle Log") == "%27Grow%20Cycle%20Log%27"


# --------------------------------------------------------------------------- #
# Round-trip: export -> .xlsx -> parse (the whole point)
# --------------------------------------------------------------------------- #
def _roundtrip(db, path) -> parse.ParsedWorkbook:
    w = writer.XlsxWriter(str(path))
    export.push(db, w)
    w.close()
    wb = load_workbook(path, read_only=True, data_only=True)
    return parse.parse_workbook(wb)


def test_export_roundtrips_through_the_parser(seeded, tmp_path):
    parsed = _roundtrip(seeded, tmp_path / "master.xlsx")

    strains = {s.name: s for s in parsed.strains}
    assert {"Stargazer", "Penis Envy"} <= set(strains)
    sg = strains["Stargazer"]
    assert sg.ease_rating == 8
    assert sg.grow_again is True
    assert sg.vendor == "Sporeworks"
    assert sg.library_status == "active"  # "Active" status cell -> active

    batches = {b.lot_code: b for b in parsed.batches}
    assert "T-01-F1" in batches
    assert batches["T-01-F1"].stage == "harvesting"
    assert batches["T-01-F1"].inoculated_on == date(2026, 5, 1)

    harvests = {h.lot_code: h for h in parsed.harvests}
    assert harvests["T-01-F1"].fresh_g == 445
    assert harvests["T-01-F1"].dry_g == 46.8

    customers = {c.name: c for c in parsed.customers}
    assert customers["Daniel Childs"].channel == "distributor"


def test_full_loop_export_then_reimport_into_a_fresh_db(seeded, tmp_path, session):
    """Export DB #1 -> sheet -> import into DB #2. The cultivation spine and
    strains survive the whole loop."""
    from backend.app.sheet.sinks import SqliteSink

    parsed = _roundtrip(seeded, tmp_path / "loop.xlsx")

    # Fresh, independent DB stands in for "the app on another machine".
    engine = create_engine(f"sqlite:///{tmp_path/'loop.db'}", future=True)
    Base.metadata.create_all(bind=engine)
    db2 = sessionmaker(bind=engine, future=True)()
    try:
        SqliteSink(db2).run(parsed)
        names = {s.name for s in db2.query(models.Strain).all()}
        assert {"Stargazer", "Penis Envy"} <= names
        h = (
            db2.query(models.Harvest)
            .join(models.Batch)
            .filter(models.Batch.lot_code == "T-01-F1")
            .one()
        )
        assert h.weight_kg == 0.445
    finally:
        db2.close()
        engine.dispose()


# --------------------------------------------------------------------------- #
# XlsxWriter behaviour
# --------------------------------------------------------------------------- #
def test_xlsx_writer_preserves_unowned_tabs(seeded, tmp_path):
    path = tmp_path / "mixed.xlsx"
    # Seed a workbook with a tab the app doesn't own.
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = "Protocols"
    wb["Protocols"]["A1"] = "keep me"
    wb.save(path)

    w = writer.XlsxWriter(str(path))
    export.push(seeded, w)
    w.close()

    back = load_workbook(path)
    assert "Protocols" in back.sheetnames          # untouched
    assert back["Protocols"]["A1"].value == "keep me"
    assert "Strain Library" in back.sheetnames      # newly written


def test_append_row_creates_tab_with_header(tmp_path):
    path = tmp_path / "append.xlsx"
    w = writer.XlsxWriter(str(path))
    spec = layout.BY_KEY["strains"]
    w.append_row(spec.tab, spec.header, ["Blue Meanie", "Active", "", None,
                                         "high", 7, "Yes", "", "note"])
    w.commit()
    w.close()

    wb = load_workbook(path)
    ws = wb["Strain Library"]
    assert [c.value for c in ws[1]][:2] == ["Strain", "Status"]
    assert ws.cell(row=2, column=1).value == "Blue Meanie"


# --------------------------------------------------------------------------- #
# Non-destructive upsert — the core "not clunky/destructive" guarantee
# --------------------------------------------------------------------------- #
def test_upsert_updates_in_place_no_duplicates(seeded, tmp_path):
    path = tmp_path / "up.xlsx"
    w = writer.XlsxWriter(str(path))
    export.push(seeded, w); w.close()

    # Change the strain in the app, push again — the row is updated, not duped.
    sg = seeded.query(models.Strain).filter_by(name="Stargazer").one()
    sg.vendor = "NewVendor"
    seeded.commit()
    w = writer.XlsxWriter(str(path))
    result = export.push(seeded, w); w.close()

    assert result["strains"]["updated"] == 2  # both strains matched & updated
    assert result["strains"]["appended"] == 0
    ws = load_workbook(path)["Strain Library"]
    names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert names.count("Stargazer") == 1                     # not duplicated
    vendor_col = ws[1].index(next(c for c in ws[1] if c.value == "Vendor")) + 1
    sg_row = names.index("Stargazer") + 2
    assert ws.cell(row=sg_row, column=vendor_col).value == "NewVendor"


def test_upsert_preserves_operator_columns_and_rows(seeded, tmp_path):
    from openpyxl import Workbook
    path = tmp_path / "op.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Strain Library"
    spec = layout.BY_KEY["strains"]
    ws.append(list(spec.header) + ["Operator Note"])
    ws.append(["Stargazer", "Active", "old", None, "", 7, "Yes", "", "", "KEEP ME"])
    ws.append(["ManualOnly", "Active", "x", None, "", 5, "Yes", "", "", "hand-added"])
    wb.save(path)

    w = writer.XlsxWriter(str(path))
    export.push(seeded, w); w.close()

    ws = load_workbook(path)["Strain Library"]
    grid = {row[0]: row for row in ws.iter_rows(values_only=True)}
    assert grid["Stargazer"][-1] == "KEEP ME"       # operator column preserved
    assert "ManualOnly" in grid                       # operator row untouched
    assert grid["ManualOnly"][-1] == "hand-added"


def test_upsert_preserves_operator_column_between_owned_columns(seeded, tmp_path):
    # An operator column inserted *between* two owned columns must survive an
    # update — we write owned cells individually, never a blanking span.
    from openpyxl import Workbook
    path = tmp_path / "mid.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Strain Library"
    hdr = ["Strain", "Status", "Operator Mid", "Vendor", "Inoculated",
           "Potency", "Ease", "Grow Again", "Tub/Bag ID", "Notes"]
    ws.append(hdr)
    ws.append(["Stargazer", "Active", "MID KEEP", "old", None, "", 7, "Yes", "", ""])
    wb.save(path)

    w = writer.XlsxWriter(str(path))
    export.push(seeded, w); w.close()

    ws = load_workbook(path)["Strain Library"]
    row = next(r for r in ws.iter_rows(values_only=True) if r[0] == "Stargazer")
    assert row[2] == "MID KEEP"        # interspersed operator column untouched
    assert row[3] == "Sporeworks"      # owned Vendor column still updated


# --------------------------------------------------------------------------- #
# Sync-state + auto-push
# --------------------------------------------------------------------------- #
def test_mark_dirty_and_pushed(seeded):
    assert state.snapshot(seeded)["dirty_count"] == 0
    state.mark_dirty(seeded, 3)
    assert state.snapshot(seeded)["dirty_count"] == 3
    state.mark_pushed(seeded)
    snap = state.snapshot(seeded)
    assert snap["dirty_count"] == 0
    assert snap["last_pushed_at"] is not None


def test_autosync_disabled_marks_dirty_without_pushing(monkeypatch, seeded, tmp_path):
    monkeypatch.delenv("SHEET_SYNC_AUTO", raising=False)
    monkeypatch.setenv("MASTER_SHEET_PATH", str(tmp_path / "should_not_exist.xlsx"))
    autosync.notify(seeded)
    assert state.snapshot(seeded)["dirty_count"] == 1
    assert not (tmp_path / "should_not_exist.xlsx").exists()  # no push happened


def test_autosync_run_push_writes_and_clears_dirty(monkeypatch, seeded, tmp_path):
    target = tmp_path / "auto.xlsx"
    monkeypatch.setenv("MASTER_SHEET_PATH", str(target))
    monkeypatch.delenv("MASTER_SHEET_GOOGLE_ID", raising=False)
    monkeypatch.delenv("MASTER_SHEET_FILE_ID", raising=False)
    state.mark_dirty(seeded, 2)

    # run_push opens its own session; point it at this test's engine.
    factory = sessionmaker(bind=seeded.get_bind(), future=True)
    counts = autosync.run_push(session_factory=factory)
    assert counts["strains"]["appended"] == 2
    assert target.exists()
    seeded.expire_all()
    assert state.snapshot(seeded)["dirty_count"] == 0


def test_autosync_run_push_noop_without_target(monkeypatch):
    for var in ("MASTER_SHEET_GOOGLE_ID", "MASTER_SHEET_FILE_ID", "MASTER_SHEET_PATH"):
        monkeypatch.delenv(var, raising=False)
    assert autosync.run_push() is None  # no target -> best-effort no-op


# --------------------------------------------------------------------------- #
# API — status / push / download over the real app
# --------------------------------------------------------------------------- #
@pytest.fixture()
def client(tmp_path):
    """A TestClient whose request sessions hit an isolated temp SQLite DB.

    Overriding the get_db dependency (rather than reloading modules) keeps the
    shared app object intact and side-effect free for the other test files.
    """
    from backend.app.database import get_db
    from backend.app.main import app

    engine = create_engine(f"sqlite:///{tmp_path/'api.db'}", future=True)
    Base.metadata.create_all(bind=engine)
    TestingSession = sessionmaker(bind=engine, autoflush=False, future=True)

    def override_get_db():
        db = TestingSession()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        with TestClient(app) as c:
            yield c
    finally:
        app.dependency_overrides.pop(get_db, None)
        engine.dispose()


def test_status_reports_no_target_when_unconfigured(client, monkeypatch):
    for var in ("MASTER_SHEET_GOOGLE_ID", "MASTER_SHEET_FILE_ID", "MASTER_SHEET_PATH"):
        monkeypatch.delenv(var, raising=False)
    r = client.get("/api/sync/status")
    assert r.status_code == 200
    body = r.json()
    assert body["write_target"]["configured"] is False
    assert body["auto_push"] is False
    assert body["sync_state"]["dirty_count"] == 0


def test_status_read_source_honors_google_id(client, monkeypatch):
    # Read source and pull must agree: a native Google Sheet id is a valid
    # source, so status must report it (not fall back to the default file).
    for var in ("MASTER_SHEET_PATH", "MASTER_SHEET_FILE_ID"):
        monkeypatch.delenv(var, raising=False)
    monkeypatch.setenv("MASTER_SHEET_GOOGLE_ID", "sheet-abc")
    body = client.get("/api/sync/status").json()
    assert body["read_source"] == {"configured": True, "kind": "google_sheet", "ref": "sheet-abc"}
    # resolve_workbook uses the same id (no silent fall-back to DEFAULT_FILE_ID).
    from backend.app.sheet import source
    assert source.resolve_read_file_id() == "sheet-abc"


def test_push_then_download_over_api(client, monkeypatch, tmp_path):
    # Create a strain through the real API…
    r = client.post("/api/strains", json={"name": "API Strain", "ease_rating": 6})
    assert r.status_code == 201

    # …push to a local .xlsx target…
    target = tmp_path / "api-master.xlsx"
    monkeypatch.setenv("MASTER_SHEET_PATH", str(target))
    monkeypatch.delenv("MASTER_SHEET_GOOGLE_ID", raising=False)
    monkeypatch.delenv("MASTER_SHEET_FILE_ID", raising=False)
    push = client.post("/api/sync/push")
    assert push.status_code == 200, push.text
    assert push.json()["written"]["strains"]["rows"] == 1
    assert target.exists()

    # Creating the strain marked the app dirty; the push cleared it.
    st = client.get("/api/sync/status").json()["sync_state"]
    assert st["dirty_count"] == 0
    assert st["last_pushed_at"] is not None

    # …and the download endpoint returns a real .xlsx too.
    dl = client.get("/api/sync/workbook.xlsx")
    assert dl.status_code == 200
    assert dl.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert dl.content[:2] == b"PK"  # xlsx is a zip
