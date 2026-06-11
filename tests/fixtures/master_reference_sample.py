"""Build a small .xlsx that mirrors the real Master Cultivation Reference
structure — banner rows, section sub-headers, and total rows included — so the
importer is exercised against the same shapes it meets in production. Rows are a
faithful subset of the live June 2026 sheet.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


def _add(ws, rows: list[list]) -> None:
    for row in rows:
        ws.append(row)


def build(path: str | Path) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    env = wb.create_sheet("Environment")
    _add(env, [
        ["GROW ROOM SETUP & ENVIRONMENT", None, None, None],
        ["Equipment", "Spec / Notes", "Status", "Last Checked"],
        ["Martha Tent", "Wire-shelf style. Relocated to basement Jun 7.", "Active", "Jun 7, 2026"],
        ["Humidifier", "Ultrasonic. Jun 10: set to 85-92% range.", "Active", "Daily"],
        ["CO2 Monitor", "Inkbird IAM-T2. Jun 9: 418 ppm.", "Active", "Jun 9, 2026"],
        [None, None, None, None],
        ["ENVIRONMENTAL TARGETS", None, None, None],
        ["Parameter", "Target", "Current", "Notes"],
        ["Temperature — Fruiting", "75-78F", "75F (Jun 10)", "Back in band."],
    ])

    sl = wb.create_sheet("Strain Library")
    _add(sl, [
        ["STRAIN LIBRARY — Active, Colonizing & Inoculating"],
        ["Strain", "Status", "Vendor", "Inoculated", "Transferred", "Potency",
         "Ease", "Grow Again", "Tub/Bag ID", "Notes / Genetics"],
        ["MEDICINAL — ACTIVE / FRUITING"],
        ["Stargazer", "Active — F3 fruiting", "Local shop", datetime(2026, 3, 30),
         datetime(2026, 5, 4), "Moderate", "8/10", "Yes", "T-01", "P. cubensis. F1 445g."],
        ["Illusion Weaver", "Active — F3 fruiting", "Local shop", datetime(2026, 3, 30),
         datetime(2026, 5, 4), "Moderate", "7/10", "Yes", "T-02", "P. cubensis. Perimeter pinning."],
        ["MEDICINAL — COLONIZING"],
        ["Golden Teacher (bag)", "Colonizing", "Sporeworks", datetime(2026, 5, 21),
         "TBD", "Moderate-High", "9/10", "TBD", "G-03", "Classic, reliable, beginner-friendly."],
        ["Natalensis", "Colonizing — fastest of batch", "-", "TBD", "TBD",
         "7-8", "TBD", "TBD", "-", "P. natalensis — South African species."],
        ["MEDICINAL — AWAITING INOCULATION"],
        ["Penis Envy", "Awaiting inoculation", "Sporeworks", datetime(2026, 5, 21),
         "TBD", "Very High", "5/10", "TBD", "-", "Mutant variety. Highest potency."],
    ])

    fr = wb.create_sheet("Fridge & Incoming")
    _add(fr, [
        ["FRIDGE SPORE INVENTORY"],
        ["Strain", "Status", "Location", "Qty", "Potency", "Ease", "Priority",
         "Vendor", "Acquired", "Notes"],
        ["Golden Teacher", "Inoculated — in bag", "Fridge", "1 syringe", "Moderate-High",
         "9/10", "****", "Sporeworks", datetime(2026, 5, 21), "In bag. Classic reliable strain."],
        ["Blue Meanie", "Inoculated — in bag", "Fridge", "0 (used Jun 9)", "High", "7/10",
         "****", "Local", "-", "Fast colonizer. High potency."],
        ["FUNCTIONAL & SPECIALTY MUSHROOMS — INCOMING (North Spore #737436)"],
        ["Strain", "Status", "Vendor", "Format", "Use / Market", "Ease", "Notes"],
        ["Blue Oyster", "En route", "North Spore", "LC x2 + spawn x3", "Restaurants + markets",
         "9/10", "Harbor Blue PO1. Fastest seller."],
        ["Lion's Mane", "En route", "North Spore", "LC x2", "Wellness market", "8/10",
         "Lion's Brain HE1. Highest value/gram."],
        ["SUPPLIES & SUBSTRATE — INCOMING (North Spore)"],
        ["Item", "Qty", "Order", None, "Notes"],
        ["'Boomr Bin' Monotubs + filters", "4", "#737436", None, ""],
    ])

    jar = wb.create_sheet("Jar Inventory")
    _add(jar, [
        ["JAR INVENTORY — Grown Dried Product on Hand"],
        ["Jar ID", "Strain", "Flush", "Harvest Date", "Dry (g)", "Used (g)",
         "Remaining (g)", "Est. Wholesale", "Est. Distributor", "Est. Retail", "Notes"],
        ["J-01", "Stargazer", "F1", datetime(2026, 5, 17), 31.2, 0, 31.2,
         "$94-$156", "$218-$250", "$374-$468", ""],
        ["J-02", "Illusion Weaver", "F1", datetime(2026, 5, 19), 46.8, 15, 31.8,
         "$95-$159", "$223-$254", "$382-$477", "Jun 5: 15g sold."],
        ["TOTAL", None, None, None, 78.0, 15, 63.0, None, None, None, "All grown jars"],
        [None],
        ["PRICING REFERENCE (Medicinal)"],
        ["Tier", "Volume", "Price / Gram", "Notes"],
        ["Wholesale", "2 lb minimum", "$3-5/g", "High volume buyers."],
        ["Distributor", "Up to 2 lbs", "$7-8/g", "Facilitators, shamans."],
        ["Retail", "Low volume", "$12-15/g", "Individual buyers."],
    ])

    sg = wb.create_sheet("Sourced Finished Goods")
    _add(sg, [
        ["SOURCED FINISHED GOODS — externally sourced product backing the menu"],
        ["These strains appear on the menu as Available but are NOT grown in-house (yet)."],
        ["Strain", "On-Hand (g)", "Used (g)", "Remaining (g)", "Incoming", "Last Updated", "Notes"],
        ["Cosmos", 112, 7, 105, "-", datetime(2026, 6, 10), "7g sold Jun 5. 85g received."],
        ["White Ape", 10, 8, 2, "-", datetime(2026, 6, 5), "OUT OF STOCK on client menu."],
    ])

    sal = wb.create_sheet("Sales Log")
    _add(sal, [
        ["SALES LOG"],
        ["Date", "Buyer", "Strain(s)", "Grams", "Amount ($)", "Tier", "Source / Notes", "Payment"],
        [datetime(2026, 6, 5), "Daniel Childs", "Illusion Weaver + Cosmos", 15, 300,
         "Retail", "8g IW (grown) + 7g Cosmos (sourced)", "Paid"],
        [datetime(2026, 6, 5), "Adam Nugent", "Tidal Wave", 10, 180, "Retail",
         "Sourced inventory. New buyer.", "Paid"],
        ["TOTAL", None, None, 25, 480, None, None, None],
        ["Collected:", 480, None, None, None, None, None, None],
        ["Outstanding:", 0, None, None, None, None, None, None],
    ])

    ht = wb.create_sheet("Harvest Tracker")
    _add(ht, [
        ["HARVEST TRACKER"],
        ["#", "Strain", "Tub", "Flush", "Harvest Date", "Fresh (g)", "Dry (g)", "Dry %", "Notes"],
        [1, "Stargazer", "T-01", "F1", datetime(2026, 5, 17), 445, 31.2, "7.0%", "Slightly early."],
        [2, "Illusion Weaver", "T-02", "F1", datetime(2026, 5, 19), 723.5, 46.8, "6.5%", "Split harvest."],
        ["TOTALS", None, None, None, None, 1168.5, 78.0, None, "Cumulative"],
    ])

    gc = wb.create_sheet("Grow Cycle Log")
    _add(gc, [
        ["GROW CYCLE LOG — one row per tub per flush"],
        ["Cycle#", "Strain", "Tub", "Flush", "Inoculated", "Transferred", "First Pins",
         "Harvest Date", "Fresh (g)", "Dry (g)", "Contam?", "Issues", "Notes"],
        [1, "Stargazer", "T-01", "F1", "Mar 30", "May 4", "~May 10-11",
         datetime(2026, 5, 17), 445, 31.2, "None", "Wet substrate", "Heavy flush."],
        [5, "BV (Bluey Vuittons)", "T-03", "F1", datetime(2026, 5, 29), "TBD", "TBD",
         "TBD", None, None, None, None, "Grain bag 1 of 3. Priority run."],
    ])

    bp = wb.create_sheet("Buyers & Pricing")
    _add(bp, [
        ["BUYER PIPELINE"],
        ["Name", "Tier", "Role", "Price/g", "Volume Est.", "Last Contact", "Follow-up", "Status", "Notes"],
        ["Jackie Brinkerhoff", "Distributor", "Shaman", "$7-8/g", "", datetime(2026, 6, 9),
         "Restock when back in town", "🟢 In contact", "Spiritual mentor."],
        ["Daniel Childs", "Retail", "Full brother", "$12-15/g", "", datetime(2026, 6, 5),
         "", "Active", "Jun 5: 8g IW + 7g Cosmos = $300."],
        ["Harmons Grocery", "Wholesale/Retail", "Local grocery", "$8-10/lb", "High", "",
         "Month 2-4", "Not contacted", "Murray or Sugar House first."],
    ])

    ven = wb.create_sheet("Vendors")
    _add(ven, [
        ["SUPPLY VENDORS"],
        ["Vendor", "Products", "URL", "Rating", "Notes"],
        ["Sporeworks", "GT, Hillbilly, PE spore syringes", "sporeworks.com", "*****", "Acquired May 21, 2026."],
        ["North Spore", "LC syringes, spawn, grain bags", "northspore.com", "*****", "Primary functional supplier."],
        ["Inkbird", "Monitoring equipment", "inkbird.com", "*****", "IBT temp/humidity + IAM-T2 CO2."],
        ["FUNCTIONAL / CHAGA SOURCING LEADS"],
        ["Vendor", "Region", "URL", "Contact Priority", "Notes"],
        ["Birch Boys", "Adirondack NY wild harvest", "birchboys.com", "1st contact", "Most credible domestic source."],
        ["CHAGA NOTE", "-", "-", "-", "Cannot be cultivated to quality — wild harvest only."],
    ])

    pr = wb.create_sheet("Protocols")
    _add(pr, [
        ["INOCULATION DAY"],
        ["[ ]", "Clean workspace thoroughly with 70% isopropyl alcohol"],
        ["[ ]", "Inject 1-2 mL per bag through self-healing port"],
        ["[ ]", "Place in dark 72-76F location"],
        ["HARVEST DAY"],
        ["[ ]", "Gloves on or thoroughly clean hands"],
        ["[ ]", "Harvest at veil break — twist grip low on stem"],
        ["[ ]", "Weigh dry, calculate ratio, log in Harvest Tracker"],
    ])

    ts = wb.create_sheet("Troubleshooting")
    _add(ts, [
        ["CONTAMINATION & MORPHOLOGY GUIDE"],
        ["Type", "Appearance", "Cause", "Action"],
        ["Green/Black Mold (Trich)", "Green or black spots on grain", "Poor sterile technique",
         "Isolate immediately. Discard if widespread."],
        ["Overlay", "Thick white mat, no pins", "Too dry surface", "Scratch lightly. Cold dunk."],
        [None],
        ["SYMPTOM -> FIX"],
        ["Symptom", "Likely Cause", "Fix"],
        ["Stalled pins 4-5 days", "CO2 buildup confirmed", "Extend tent opening. Bump FAE."],
        ["Long skinny stems", "High CO2 / low FAE", "Check IAM-T2. Increase FAE."],
        [None],
        ["INCIDENT LOG"],
        ["Date", "Issue", "Root Cause", "Resolution"],
        ["May 17", "F2 Stargazer pins damaged", "Misidentified healthy pins as aborts",
         "When F2 pins visible: mist and return."],
        ["May 21", "JMF colonization stalled — 87F", "Heating pad hot; checked pad not bag",
         "Moved to ambient 76F. Verify AT THE BAG."],
    ])

    path = str(path)
    wb.save(path)
    return path


if __name__ == "__main__":
    out = build("master_reference_sample.xlsx")
    print(f"wrote {out}")
